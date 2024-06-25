# test the excel sheet to print data
import openpyxl as xl

import pandas as pd

#file path
file_path='Krishan Mac Remove 13-6-2024_update2 (1)_ef38bfd8-34c2-49fe-b267-9bf8df10d20c.xlsx'

sheet = xl.load_workbook(file_path)

macSheet = sheet.active
macSheet = sheet['Sheet1']

print(macSheet)

for r in range(2,macSheet.max_row+1):
    value = [macSheet.cell(row=r,column=i).value for i in range(3,5)]
    print(value)


print("Total number of rows: "+str(macSheet.max_row)+" Total number of columns: "+str(macSheet.max_column))

print(type(value))


# dataframe = datafile.active

#for r in range(0, dataframe.max_row):
#    for c in dataframe.iter_cols(1, dataframe.max_column):
#        print(c[r].value)
#
