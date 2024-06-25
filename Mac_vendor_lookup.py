# import openpyxl, pandas and mac_vendor_lookup library.
import openpyxl as xl
import pandas as pd
from mac_vendor_lookup import MacLookup

maclook = MacLookup()
path = 'Surinder_ mac_sheet.xlsx'

new_save = 'Surinder_ mac_sheet_vendor_list.xlsx'

def mac_vendor_finder(mac):
    result = str("")
    try:
        result = maclook.lookup(mac) 
    except:
        result = "none"

    return result
    
def main():

    #file path
    file_path = path

    # open exsiting mac data excel sheet
    sheet1 = xl.load_workbook(file_path)

    macSheet = sheet1.active
    macSheet = sheet1['Sheet1']

    # Call a Workbook() function of openpyxl  
    # to create a new blank Workbook object 
    wb = xl.Workbook() 
    sheet2 = wb.active
    
    # Sheets can be added to workbook with the 
    # workbook object's create_sheet() method.  
    # wb.create_sheet(index = 1 , title = "Mac_filter_Sheet") 
    # the save() workbook method. 
    wb.save(new_save) 

    #vendor_list = list([])
    for r in range(1,macSheet.max_row+1):
        value = [macSheet.cell(row=r,column=i).value for i in range(1,macSheet.max_column+1)]
        getVendor1 = mac_vendor_finder(str(value[2]))
        getVendor2 = mac_vendor_finder(str(value[3]))

        vendor_list = list([value[0], value[1], value[2], getVendor1, value[3], getVendor2, value[4], value[5], value[6], value[7], value[8]])
        
        # add data in new sheet
        sheet2.append(vendor_list)
        # print(vendor_list)

    

    # Anytime you modify the Workbook object 
    # or its sheets and cells, the spreadsheet 
    # file will not be saved until you call 
    # the save() workbook method. 
    wb.save(new_save) 

    #print("Total number of rows: "+str(macSheet.max_row)+" Total number of columns: "+str(macSheet.max_column))
    #print(type(value))
        

if __name__ == "__main__":
    main()
